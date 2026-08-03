"""
PDF & Document Extractor Service
Extracts text, section headings, and metadata from uploaded policy & procedure documents.
Supports PDF (via PyMuPDF/PyPDF), Excel (.xlsx/.xls/.csv), and text files.
"""

import os
import re
from typing import Dict, Any, List

class PDFExtractor:
    @staticmethod
    def extract_text_from_file(file_path: str) -> Dict[str, Any]:
        """
        Reads text from PDF, Excel, or TXT document and returns structured content with metadata.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        file_name = os.path.basename(file_path)

        if ext == ".pdf":
            return PDFExtractor._extract_pdf(file_path, file_name)
        elif ext in [".xlsx", ".xls"]:
            return PDFExtractor._extract_excel(file_path, file_name)
        elif ext in [".txt", ".md", ".json", ".csv", ".log"]:
            return PDFExtractor._extract_text(file_path, file_name)
        else:
            # Fallback text extraction
            return PDFExtractor._extract_text(file_path, file_name)

    @staticmethod
    def _extract_pdf(file_path: str, file_name: str) -> Dict[str, Any]:
        full_text = ""
        pages_content = []
        num_pages = 1

        # 1. Try PyMuPDF (fitz) first for maximum accuracy
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(file_path)
            num_pages = len(doc)
            for idx, page in enumerate(doc):
                text = page.get_text() or ""
                pages_content.append({"page_number": idx + 1, "content": text})
                full_text += f"\n--- Page {idx + 1} ---\n" + text
            doc.close()
        except Exception:
            # 2. Fallback to PyPDF
            try:
                from pypdf import PdfReader
                reader = PdfReader(file_path)
                num_pages = len(reader.pages)
                for idx, page in enumerate(reader.pages):
                    text = page.extract_text() or ""
                    pages_content.append({"page_number": idx + 1, "content": text})
                    full_text += f"\n--- Page {idx + 1} ---\n" + text
            except Exception as e:
                full_text = f"Error extracting PDF text: {e}"

        headings = PDFExtractor._detect_headings(full_text)

        return {
            "filename": file_name,
            "file_type": "PDF",
            "page_count": num_pages,
            "char_count": len(full_text),
            "word_count": len(full_text.split()),
            "full_text": full_text,
            "pages": pages_content,
            "detected_headings": headings
        }

    @staticmethod
    def _extract_excel(file_path: str, file_name: str) -> Dict[str, Any]:
        """Extracts text from Excel spreadsheets (.xlsx, .xls)"""
        full_text = ""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                full_text += f"\n--- Sheet: {sheetname} ---\n"
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(val) for val in row if val is not None]
                    if row_vals:
                        full_text += " | ".join(row_vals) + "\n"
        except Exception as e:
            full_text = f"Error reading Excel file: {e}"

        headings = PDFExtractor._detect_headings(full_text)

        return {
            "filename": file_name,
            "file_type": "Excel Spreadsheet",
            "page_count": 1,
            "char_count": len(full_text),
            "word_count": len(full_text.split()),
            "full_text": full_text,
            "pages": [{"page_number": 1, "content": full_text}],
            "detected_headings": headings
        }

    @staticmethod
    def _extract_text(file_path: str, file_name: str) -> Dict[str, Any]:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()

        headings = PDFExtractor._detect_headings(text)

        return {
            "filename": file_name,
            "file_type": "Text/Document",
            "page_count": max(1, len(text) // 2500),
            "char_count": len(text),
            "word_count": len(text.split()),
            "full_text": text,
            "pages": [{"page_number": 1, "content": text}],
            "detected_headings": headings
        }

    @staticmethod
    def _detect_headings(text: str) -> List[str]:
        heading_patterns = [
            r'^(?:Clause|Section|Policy|Procedure|Module)\s+\d+(?:\.\d+)*\b.*$',
            r'^\d+\.\d+(?:\.\d+)*\s+[A-Z].*$',
            r'^[A-Z0-9\s\-\:]{4,50}$'
        ]
        
        found = []
        lines = text.splitlines()
        for line in lines:
            line_str = line.strip()
            if 3 < len(line_str) < 80:
                for pattern in heading_patterns:
                    if re.match(pattern, line_str, re.IGNORECASE):
                        found.append(line_str)
                        break
            if len(found) >= 15:
                break

        return found
