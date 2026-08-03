"""
Report Generator Service
Generates executive-ready PDF and Excel audit reports.
"""

import os
from datetime import datetime
from typing import Dict, Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ReportGenerator:
    @staticmethod
    def generate_pdf_report(audit_data: Dict[str, Any], output_dir: str) -> str:
        """
        Creates a PDF audit report from audit results and saves it to output_dir.
        Returns the absolute filepath of the generated PDF.
        """
        os.makedirs(output_dir, exist_ok=True)
        audit_id = audit_data.get("audit_id", "AUD-001")
        filename = f"ISO_Audit_Report_{audit_id}.pdf"
        filepath = os.path.join(output_dir, filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()

        # Custom Palette & Styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#0F172A'),
            spaceAfter=6
        )

        subtitle_style = ParagraphStyle(
            'DocSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=11,
            leading=14,
            textColor=colors.HexColor('#475569'),
            spaceAfter=15
        )

        h2_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=18,
            textColor=colors.HexColor('#1E293B'),
            spaceBefore=12,
            spaceAfter=8
        )

        body_style = ParagraphStyle(
            'BodyTextCustom',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9.5,
            leading=13,
            textColor=colors.HexColor('#334155')
        )

        table_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=11,
            textColor=colors.white
        )

        elements = []

        # Header Title
        elements.append(Paragraph("ISO COMPLIANCE AUDIT REPORT", title_style))
        elements.append(Paragraph(
            f"Generated on {datetime.now().strftime('%B %d, %Y')} | Audit Reference: <b>{audit_id}</b>",
            subtitle_style
        ))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#3B82F6'), spaceAfter=15))

        # Executive Summary Box
        score = audit_data.get("overall_score", 0)
        risk = audit_data.get("risk_rating", "MEDIUM")
        standards_str = ", ".join(audit_data.get("standards", ["ISO 9001"]))

        score_color = "#10B981" if score >= 80 else "#F59E0B" if score >= 60 else "#EF4444"

        summary_data = [
            [
                Paragraph(f"<b>Document Audited:</b> {audit_data.get('filename')}", body_style),
                Paragraph(f"<b>Overall Score:</b> <font color='{score_color}'><b>{score}%</b></font>", body_style)
            ],
            [
                Paragraph(f"<b>Standards:</b> {standards_str}", body_style),
                Paragraph(f"<b>Risk Rating:</b> <b>{risk}</b>", body_style)
            ],
            [
                Paragraph(f"<b>Total Clauses Audited:</b> {audit_data.get('total_clauses_audited', 0)}", body_style),
                Paragraph(f"<b>Compliant:</b> {audit_data.get('compliant_count', 0)} | <b>Minor NC:</b> {audit_data.get('minor_nc_count', 0)} | <b>Major NC:</b> {audit_data.get('major_nc_count', 0)}", body_style)
            ]
        ]

        summary_table = Table(summary_data, colWidths=[3.6 * inch, 3.6 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#F1F5F9')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 15))

        # Executive Summary Narrative
        elements.append(Paragraph("Executive Narrative", h2_style))
        elements.append(Paragraph(audit_data.get("executive_summary", "Audit completed successfully."), body_style))
        elements.append(Spacer(1, 15))

        # Audit Findings Breakdown Table
        elements.append(Paragraph("Detailed Risk & Clause Compliance Assessment", h2_style))

        findings = audit_data.get("findings", [])
        table_rows = [
            [
                Paragraph("Clause & Dept", table_header_style),
                Paragraph("Status", table_header_style),
                Paragraph("Risk (L x S = R)", table_header_style),
                Paragraph("Impact", table_header_style),
                Paragraph("Evidence & Findings", table_header_style),
                Paragraph("Recommendations", table_header_style)
            ]
        ]

        for item in findings:
            status_text = item.get("status", "COMPLIANT").replace("_", " ")
            status_fg = "#166534" if "COMPLIANT" in item.get("status") else "#92400E" if "MINOR" in item.get("status") else "#991B1B"
            status_p = Paragraph(f"<font color='{status_fg}'><b>{status_text}</b></font>", body_style)

            # Extract or compute Risk Matrix parameters (L, S, R, Impact)
            score_val = item.get("score", 75)
            l_val = item.get("likelihood", 3)
            s_val = item.get("severity", 3)
            r_val = item.get("risk_score", l_val * s_val)
            impact_val = item.get("impact", "Medium" if r_val >= 8 else "Low")
            impact_fg = "#EF4444" if impact_val == "High" else "#F59E0B" if impact_val == "Medium" else "#10B981"
            impact_p = Paragraph(f"<font color='{impact_fg}'><b>{impact_val}</b></font>", body_style)

            dept_val = item.get("department", "Operations")

            table_rows.append([
                Paragraph(f"<b>{item.get('clause_id')}</b><br/><font color='#64748B'>{dept_val}</font>", body_style),
                status_p,
                Paragraph(f"L:{l_val} x S:{s_val}<br/><b>R = {r_val}</b>", body_style),
                impact_p,
                Paragraph(f"<b>Evidence:</b> {item.get('evidence_found')}<br/><b>Gaps:</b> {item.get('gaps_identified')}", body_style),
                Paragraph(item.get('recommendations', ''), body_style)
            ])

        findings_table = Table(table_rows, colWidths=[1.1 * inch, 1.1 * inch, 0.9 * inch, 0.7 * inch, 1.8 * inch, 1.6 * inch])
        findings_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (2,0), (3,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))


        elements.append(findings_table)
        elements.append(Spacer(1, 20))

        # Sign-off block
        elements.append(KeepTogether([
            Paragraph("Audit Certification & Sign-off", h2_style),
            Paragraph("This audit report was compiled and verified using AI-Assisted ISO Compliance Auditor. Non-conformities listed herein require formal Corrective Action Requests (CAR) to be closed prior to surveillance audit certification.", body_style),
            Spacer(1, 15),
            Table([
                [Paragraph("<b>Lead Auditor Signature:</b> ___________________________", body_style), Paragraph("<b>Quality Lead Approval:</b> ___________________________", body_style)],
                [Paragraph("<b>Date:</b> " + datetime.now().strftime("%Y-%m-%d"), body_style), Paragraph("<b>Status:</b> Official Report Issued", body_style)]
            ], colWidths=[3.6 * inch, 3.6 * inch])
        ]))

        doc.build(elements)
        return filepath

    @staticmethod
    def generate_excel_report(audit_data: Dict[str, Any], output_dir: str) -> str:
        """
        Creates an Excel (.xlsx) audit report with a Summary sheet and a Findings sheet.
        Returns the absolute filepath of the generated Excel file.
        """
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

        os.makedirs(output_dir, exist_ok=True)
        audit_id = audit_data.get("audit_id", "AUD-001")
        filename = f"ISO_Audit_Report_{audit_id}.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()

        # ─── Shared Styles ───────────────────────────────────────────
        header_fill   = PatternFill("solid", fgColor="1E293B")
        header_font   = Font(bold=True, color="FFFFFF", size=11)
        title_font    = Font(bold=True, color="0F172A", size=16)
        subtitle_font = Font(italic=True, color="475569", size=10)
        label_font    = Font(bold=True, color="1E293B", size=10)
        value_font    = Font(color="334155", size=10)
        center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align    = Alignment(horizontal="left", vertical="top", wrap_text=True)

        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        STATUS_FILLS = {
            "COMPLIANT":                PatternFill("solid", fgColor="DCFCE7"),
            "MINOR_NON_CONFORMITY":     PatternFill("solid", fgColor="FEF3C7"),
            "MAJOR_NON_CONFORMITY":     PatternFill("solid", fgColor="FEE2E2"),
            "OPPORTUNITY_FOR_IMPROVEMENT": PatternFill("solid", fgColor="DBEAFE"),
        }
        STATUS_FONTS = {
            "COMPLIANT":                Font(bold=True, color="166534", size=10),
            "MINOR_NON_CONFORMITY":     Font(bold=True, color="92400E", size=10),
            "MAJOR_NON_CONFORMITY":     Font(bold=True, color="991B1B", size=10),
            "OPPORTUNITY_FOR_IMPROVEMENT": Font(bold=True, color="1E40AF", size=10),
        }

        score = audit_data.get("overall_score", 0)
        risk  = audit_data.get("risk_rating", "MEDIUM")

        # ─── SHEET 1: Executive Summary ───────────────────────────────
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 52

        def write_row(ws, row, label, value, label_fill_hex=None):
            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=value)
            lc.font = label_font
            vc.font = value_font
            lc.alignment = left_align
            vc.alignment = left_align
            lc.border = border
            vc.border = border
            if label_fill_hex:
                lc.fill = PatternFill("solid", fgColor=label_fill_hex)
                vc.fill = PatternFill("solid", fgColor=label_fill_hex)
            ws.row_dimensions[row].height = 22

        # Title block
        ws1.merge_cells("A1:B1")
        t = ws1["A1"]
        t.value = "ISO COMPLIANCE AUDIT REPORT"
        t.font = title_font
        t.alignment = center_align
        t.fill = PatternFill("solid", fgColor="EFF6FF")
        ws1.row_dimensions[1].height = 36

        ws1.merge_cells("A2:B2")
        s = ws1["A2"]
        s.value = f"Audit Reference: {audit_id}  |  Generated: {datetime.now().strftime('%B %d, %Y')}"
        s.font = subtitle_font
        s.alignment = center_align
        ws1.row_dimensions[2].height = 18

        ws1.row_dimensions[3].height = 8  # spacer

        write_row(ws1, 4,  "Document Audited",       audit_data.get("filename", ""), "F8FAFC")
        write_row(ws1, 5,  "Standards Audited",       ", ".join(audit_data.get("standards", [])), "F8FAFC")
        write_row(ws1, 6,  "Overall Compliance Score", f"{score}%", "F8FAFC")
        write_row(ws1, 7,  "Risk Rating",              risk, "F8FAFC")
        write_row(ws1, 8,  "Total Clauses Audited",   audit_data.get("total_clauses_audited", 0), "F8FAFC")
        write_row(ws1, 9,  "Compliant Clauses",        audit_data.get("compliant_count", 0), "F8FAFC")
        write_row(ws1, 10, "Minor Non-Conformities",   audit_data.get("minor_nc_count", 0), "F8FAFC")
        write_row(ws1, 11, "Major Non-Conformities",   audit_data.get("major_nc_count", 0), "F8FAFC")
        write_row(ws1, 12, "Opportunities for Improvement", audit_data.get("ofi_count", 0), "F8FAFC")

        ws1.row_dimensions[13].height = 8  # spacer

        # Score colour-code cell
        score_cell = ws1["B6"]
        if score >= 80:
            score_cell.font = Font(bold=True, color="166534", size=12)
            score_cell.fill = PatternFill("solid", fgColor="DCFCE7")
        elif score >= 60:
            score_cell.font = Font(bold=True, color="92400E", size=12)
            score_cell.fill = PatternFill("solid", fgColor="FEF3C7")
        else:
            score_cell.font = Font(bold=True, color="991B1B", size=12)
            score_cell.fill = PatternFill("solid", fgColor="FEE2E2")

        # Executive Summary narrative
        ws1.merge_cells("A14:B14")
        h = ws1["A14"]
        h.value = "Executive Summary"
        h.font = Font(bold=True, color="FFFFFF", size=12)
        h.fill = PatternFill("solid", fgColor="1E293B")
        h.alignment = center_align
        ws1.row_dimensions[14].height = 24

        ws1.merge_cells("A15:B15")
        n = ws1["A15"]
        n.value = audit_data.get("executive_summary", "")
        n.font = value_font
        n.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws1.row_dimensions[15].height = 80

        # ─── SHEET 2: Detailed Clause Findings ───────────────────────
        ws2 = wb.create_sheet("Clause Findings")
        ws2.sheet_view.showGridLines = False

        col_widths = [12, 32, 24, 12, 42, 42, 50]
        col_names  = ["Clause ID", "Clause Name", "Title", "Score (%)", "Evidence Found", "Gaps Identified", "Recommendations"]
        for i, (w, name) in enumerate(zip(col_widths, col_names), start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w
            c = ws2.cell(row=1, column=i, value=name)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = border
        ws2.row_dimensions[1].height = 28

        for row_idx, item in enumerate(audit_data.get("findings", []), start=2):
            status = item.get("status", "COMPLIANT")
            row_data = [
                item.get("clause_id", ""),
                item.get("clause_name", ""),
                item.get("title", ""),
                item.get("score", 0),
                item.get("evidence_found", ""),
                item.get("gaps_identified", ""),
                item.get("recommendations", ""),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                c = ws2.cell(row=row_idx, column=col_idx, value=val)
                c.alignment = left_align
                c.border = border
                c.font = value_font

            # Colour-code the Status — insert in column 4 (after title)
            # We'll add a Status column after "Title" (shift Score to col 5)
            # Actually reorder: use col 4 as status
            status_cell = ws2.cell(row=row_idx, column=4, value=item.get("score", 0))
            status_cell.fill = STATUS_FILLS.get(status, PatternFill())

            ws2.row_dimensions[row_idx].height = 60

        # Add a Status column as column 4 header
        ws2.insert_cols(4)
        ws2.column_dimensions["D"].width = 24
        ws2.cell(row=1, column=4, value="Status").font = header_font
        ws2.cell(row=1, column=4).fill = header_fill
        ws2.cell(row=1, column=4).alignment = center_align
        ws2.cell(row=1, column=4).border = border

        for row_idx, item in enumerate(audit_data.get("findings", []), start=2):
            status = item.get("status", "COMPLIANT")
            sc = ws2.cell(row=row_idx, column=4, value=status.replace("_", " "))
            sc.fill = STATUS_FILLS.get(status, PatternFill())
            sc.font = STATUS_FONTS.get(status, value_font)
            sc.alignment = center_align
            sc.border = border

        # Freeze top header row on both sheets
        ws1.freeze_panes = "A4"
        ws2.freeze_panes = "A2"

        wb.save(filepath)
        return filepath


import os
from datetime import datetime
from typing import Dict, Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)

class ReportGenerator:
    @staticmethod
    def generate_pdf_report(audit_data: Dict[str, Any], output_dir: str) -> str:
        """
        Creates a PDF audit report from audit results and saves it to output_dir.
        Returns the absolute filepath of the generated PDF.
        """
        os.makedirs(output_dir, exist_ok=True)
        audit_id = audit_data.get("audit_id", "AUD-001")
        filename = f"ISO_Audit_Report_{audit_id}.pdf"
        filepath = os.path.join(output_dir, filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()

        # Custom Palette & Styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#0F172A'),
            spaceAfter=6
        )

        subtitle_style = ParagraphStyle(
            'DocSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=11,
            leading=14,
            textColor=colors.HexColor('#475569'),
            spaceAfter=15
        )

        h2_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=18,
            textColor=colors.HexColor('#1E293B'),
            spaceBefore=12,
            spaceAfter=8
        )

        body_style = ParagraphStyle(
            'BodyTextCustom',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9.5,
            leading=13,
            textColor=colors.HexColor('#334155')
        )

        table_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=11,
            textColor=colors.white
        )

        elements = []

        # Header Title
        elements.append(Paragraph("ISO COMPLIANCE AUDIT REPORT", title_style))
        elements.append(Paragraph(
            f"Generated on {datetime.now().strftime('%B %d, %Y')} | Audit Reference: <b>{audit_id}</b>",
            subtitle_style
        ))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#3B82F6'), spaceAfter=15))

        # Executive Summary Box
        score = audit_data.get("overall_score", 0)
        risk = audit_data.get("risk_rating", "MEDIUM")
        standards_str = ", ".join(audit_data.get("standards", ["ISO 9001"]))

        score_color = "#10B981" if score >= 80 else "#F59E0B" if score >= 60 else "#EF4444"

        summary_data = [
            [
                Paragraph(f"<b>Document Audited:</b> {audit_data.get('filename')}", body_style),
                Paragraph(f"<b>Overall Score:</b> <font color='{score_color}'><b>{score}%</b></font>", body_style)
            ],
            [
                Paragraph(f"<b>Standards:</b> {standards_str}", body_style),
                Paragraph(f"<b>Risk Rating:</b> <b>{risk}</b>", body_style)
            ],
            [
                Paragraph(f"<b>Total Clauses Audited:</b> {audit_data.get('total_clauses_audited', 0)}", body_style),
                Paragraph(f"<b>Compliant:</b> {audit_data.get('compliant_count', 0)} | <b>Minor NC:</b> {audit_data.get('minor_nc_count', 0)} | <b>Major NC:</b> {audit_data.get('major_nc_count', 0)}", body_style)
            ]
        ]

        summary_table = Table(summary_data, colWidths=[3.6 * inch, 3.6 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#F1F5F9')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 15))

        # Executive Summary Narrative
        elements.append(Paragraph("Executive Narrative", h2_style))
        elements.append(Paragraph(audit_data.get("executive_summary", "Audit completed successfully."), body_style))
        elements.append(Spacer(1, 15))

        # Audit Findings Breakdown Table
        elements.append(Paragraph("Clause Compliance Breakdown", h2_style))

        findings = audit_data.get("findings", [])
        table_rows = [
            [
                Paragraph("Clause", table_header_style),
                Paragraph("Status", table_header_style),
                Paragraph("Score", table_header_style),
                Paragraph("Evidence & Findings", table_header_style),
                Paragraph("Recommendations", table_header_style)
            ]
        ]

        for item in findings:
            status_text = item.get("status", "COMPLIANT").replace("_", " ")
            status_bg = "#DCFCE7" if "COMPLIANT" in item.get("status") else "#FEF3C7" if "MINOR" in item.get("status") else "#FEE2E2"
            status_fg = "#166534" if "COMPLIANT" in item.get("status") else "#92400E" if "MINOR" in item.get("status") else "#991B1B"

            status_p = Paragraph(f"<font color='{status_fg}'><b>{status_text}</b></font>", body_style)

            table_rows.append([
                Paragraph(f"<b>{item.get('clause_id')}</b><br/>{item.get('title', '')}", body_style),
                status_p,
                Paragraph(f"<b>{item.get('score')}%</b>", body_style),
                Paragraph(f"<b>Evidence:</b> {item.get('evidence_found')}<br/><b>Gaps:</b> {item.get('gaps_identified')}", body_style),
                Paragraph(item.get('recommendations', ''), body_style)
            ])

        findings_table = Table(table_rows, colWidths=[1.1 * inch, 1.2 * inch, 0.6 * inch, 2.3 * inch, 2.0 * inch])
        findings_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (2,0), (2,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))

        elements.append(findings_table)
        elements.append(Spacer(1, 20))

        # Sign-off block
        elements.append(KeepTogether([
            Paragraph("Audit Certification & Sign-off", h2_style),
            Paragraph("This audit report was compiled and verified using AI-Assisted ISO Compliance Auditor. Non-conformities listed herein require formal Corrective Action Requests (CAR) to be closed prior to surveillance audit certification.", body_style),
            Spacer(1, 15),
            Table([
                [Paragraph("<b>Lead Auditor Signature:</b> ___________________________", body_style), Paragraph("<b>Quality Lead Approval:</b> ___________________________", body_style)],
                [Paragraph("<b>Date:</b> " + datetime.now().strftime("%Y-%m-%d"), body_style), Paragraph("<b>Status:</b> Official Report Issued", body_style)]
            ], colWidths=[3.6 * inch, 3.6 * inch])
        ]))

        doc.build(elements)
        return filepath
